"""
Phase v0.2 - CSV/Excel Importer Module
Handles CSV and Excel BOM imports with validation
"""

from abc import ABC, abstractmethod
from typing import List, Dict, Any
import csv
import io

from app.schemas import Part, BOM


class BOMImporter(ABC):
    """Abstract base class for BOM importers"""
    
    @abstractmethod
    def import_bom(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Import BOM from file content
        
        Args:
            file_content: Raw file bytes
            filename: Name of the file
            
        Returns:
            Dictionary with 'bom', 'parts', 'errors', 'validation_summary'
        """
        pass
    
    @abstractmethod
    def validate_format(self, data: List[Dict[str, str]]) -> tuple[bool, List[str]]:
        """Validate data format and return (is_valid, error_list)"""
        pass


class CSVBOMImporter(BOMImporter):
    """Importer for CSV format BOMs"""
    
    REQUIRED_COLUMNS = {"part_number", "description", "quantity", "supplier"}
    OPTIONAL_COLUMNS = {"model", "category", "lead_time_weeks", "cost", "alternates"}
    
    def import_bom(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Import BOM from CSV file
        
        Expected CSV columns:
        - part_number (required)
        - description (required)
        - quantity (required)
        - supplier (required)
        - model (optional)
        - category (optional)
        - lead_time_weeks (optional)
        - cost (optional)
        - alternates (optional, semicolon-separated)
        """
        try:
            # Decode file content
            text_content = file_content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text_content))
            
            if reader.fieldnames is None:
                return {
                    "success": False,
                    "error": "CSV file is empty",
                    "bom": None,
                    "parts": [],
                    "validation_summary": {"total_rows": 0, "valid_rows": 0, "errors": 0}
                }
            
            data = list(reader)
            is_valid, errors = self.validate_format(data)
            
            if not is_valid:
                return {
                    "success": False,
                    "error": "CSV format validation failed",
                    "validation_errors": errors,
                    "bom": None,
                    "parts": [],
                    "validation_summary": {"total_rows": len(data)}
                }
            
            # Parse parts
            parts = []
            parse_errors = []
            
            for idx, row in enumerate(data, start=2):  # Start at 2 (header is row 1)
                try:
                    part = self._parse_part_row(row)
                    parts.append(part)
                except ValueError as e:
                    parse_errors.append(f"Row {idx}: {str(e)}")
            
            # Create BOM
            bom = BOM(
                name=filename.replace('.csv', ''),
                parts=parts,
                version="1.0",
                created_from_import=True
            )
            
            return {
                "success": len(parse_errors) == 0 or len(parts) > 0,
                "bom": bom,
                "parts": parts,
                "errors": parse_errors,
                "validation_summary": {
                    "total_rows": len(data),
                    "valid_rows": len(parts),
                    "error_rows": len(parse_errors)
                }
            }
        
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to import CSV: {str(e)}",
                "bom": None,
                "parts": [],
                "validation_summary": {"error": str(e)}
            }
    
    def validate_format(self, data: List[Dict[str, str]]) -> tuple[bool, List[str]]:
        """Validate CSV data has required columns"""
        if not data:
            return False, ["CSV file is empty"]
        
        first_row = data[0]
        missing_cols = self.REQUIRED_COLUMNS - set(first_row.keys())
        
        if missing_cols:
            return False, [f"Missing required columns: {', '.join(missing_cols)}"]
        
        return True, []
    
    def _parse_part_row(self, row: Dict[str, str]) -> Dict[str, Any]:
        """Parse a CSV row into a Part object"""
        part_number = row.get("part_number", "").strip()
        
        if not part_number:
            raise ValueError("part_number is required")
        
        quantity_str = row.get("quantity", "1").strip()
        try:
            quantity = int(quantity_str)
        except ValueError:
            raise ValueError(f"Invalid quantity: {quantity_str}")
        
        return {
            "part_number": part_number,
            "description": row.get("description", "").strip(),
            "quantity": quantity,
            "supplier": row.get("supplier", "").strip(),
            "model": row.get("model", "").strip() or None,
            "category": row.get("category", "").strip() or None,
            "lead_time_weeks": int(row.get("lead_time_weeks", 0)) if row.get("lead_time_weeks") else None,
            "cost": float(row.get("cost", 0)) if row.get("cost") else None,
            "alternates": [alt.strip() for alt in row.get("alternates", "").split(";") if alt.strip()]
        }


class ExcelBOMImporter(BOMImporter):
    """Importer for Excel format BOMs"""
    
    SUPPORTED_FORMATS = {".xlsx", ".xls"}
    
    def import_bom(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Import BOM from Excel file
        
        Requires openpyxl or xlrd library
        """
        try:
            import openpyxl
        except ImportError:
            return {
                "success": False,
                "error": "openpyxl library not installed. Install with: pip install openpyxl",
                "bom": None,
                "parts": [],
                "validation_summary": {}
            }
        
        try:
            from openpyxl import load_workbook
            
            # Load workbook from bytes
            workbook = load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active
            
            # Convert Excel rows to CSV-like format
            data = []
            headers = None
            
            for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if idx == 1:
                    headers = [h for h in row if h is not None]
                else:
                    row_dict = {}
                    for i, value in enumerate(row[:len(headers)]):
                        if headers[i]:
                            row_dict[headers[i].lower().replace(" ", "_")] = str(value) if value else ""
                    if any(row_dict.values()):  # Skip empty rows
                        data.append(row_dict)
            
            if not headers:
                return {
                    "success": False,
                    "error": "Excel file is empty",
                    "bom": None,
                    "parts": [],
                    "validation_summary": {}
                }
            
            # Use CSV importer logic
            csv_importer = CSVBOMImporter()
            is_valid, errors = csv_importer.validate_format(data)
            
            if not is_valid:
                return {
                    "success": False,
                    "error": "Excel format validation failed",
                    "validation_errors": errors,
                    "bom": None,
                    "parts": [],
                    "validation_summary": {"total_rows": len(data)}
                }
            
            # Parse parts
            parts = []
            parse_errors = []
            
            for idx, row in enumerate(data, start=2):
                try:
                    part = csv_importer._parse_part_row(row)
                    parts.append(part)
                except ValueError as e:
                    parse_errors.append(f"Row {idx}: {str(e)}")
            
            # Create BOM
            bom = BOM(
                name=filename.replace('.xlsx', '').replace('.xls', ''),
                parts=parts,
                version="1.0",
                created_from_import=True
            )
            
            return {
                "success": len(parse_errors) == 0 or len(parts) > 0,
                "bom": bom,
                "parts": parts,
                "errors": parse_errors,
                "validation_summary": {
                    "total_rows": len(data),
                    "valid_rows": len(parts),
                    "error_rows": len(parse_errors)
                }
            }
        
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to import Excel: {str(e)}",
                "bom": None,
                "parts": [],
                "validation_summary": {"error": str(e)}
            }
    
    def validate_format(self, data: List[Dict[str, str]]) -> tuple[bool, List[str]]:
        """Validate Excel data"""
        csv_importer = CSVBOMImporter()
        return csv_importer.validate_format(data)


def create_importer(filename: str) -> BOMImporter:
    """Factory function to create appropriate importer based on file extension"""
    extension = "." + filename.split(".")[-1].lower() if "." in filename else ""
    
    if extension == ".csv":
        return CSVBOMImporter()
    elif extension in {".xlsx", ".xls"}:
        return ExcelBOMImporter()
    else:
        raise ValueError(f"Unsupported file format: {extension}")
